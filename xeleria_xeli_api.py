"""Rutas tenantizadas para Preguntas ML y Gestión -> Entrenamiento Xeli.

No importa ``xeleria_app`` para evitar dependencias circulares. El instalador
recibe el diccionario de símbolos del módulo principal y publica el poller en
ese mismo diccionario para que lo ejecute el dispatcher multi-tenant existente.
"""

from __future__ import annotations

import io
import json
import re
import time
import unicodedata
import uuid
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Mapping, Optional, Tuple
from zoneinfo import ZoneInfo

import requests
from fastapi import File, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import PlainTextResponse, Response
from pydantic import BaseModel, Field

from xeleria_correo import (
    XeleriaCorreoError,
    normalize_postal_code,
    quote_for_model,
    shipping_dimensions_from_ml_item,
)
from xeleria_tenant_db import DEFAULT_TENANT_ID, get_effective_tenant_id
from xeleria_xeli_core import (
    XELI_CORE_VERSION,
    compose_answer,
    filter_publications_by_title,
    json_message_payload,
    knowledge_record,
    merged_settings,
    normalize_operational_table,
    normalize_response_mode,
    operational_safe_json,
    order_knowledge,
    publication_live_context,
    search_inventory_catalog,
    validate_catalog_recommendations,
)


XELI_API_VERSION = "0.4.3-shipping-policy"
_CORE: Dict[str, Any] = {}

_XELI_CANDIDATE_LIMIT = 500
_XELI_CANDIDATE_LISTING_LIMIT = 2000
_XELI_SEARCH_TERM_LIMIT = 12
_XELI_TOOL_MAX_ROUNDS = 5

_XELI_SEARCH_TOOL = {
    "type": "function",
    "function": {
        "name": "search_xeleria_inventory",
        "description": (
            "Busca en TODO el inventario tenantizado y sus publicaciones de Mercado Libre. "
            "Usala antes de afirmar disponibilidad o recomendar otro producto, variante o medida. "
            "Para expresiones ambiguas, lee las reglas del skill y busca todos los significados "
            "concretos que esas reglas enumeran; no uses solamente la palabra ambigua del comprador."
        ),
        "strict": True,
        "parameters": {
            "type": "object",
            "properties": {
                "product_terms": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Nombres o sinonimos alternativos del producto; alcanza con que coincida uno.",
                },
                "context_terms": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Palabras de contexto que mejoran el orden, por ejemplo pool o billar.",
                },
                "minimum_size_mm": {
                    "type": ["number", "null"],
                    "description": "Medida minima exclusiva. Usar para el tamano disponible inmediato superior.",
                },
                "exact_size_mm": {
                    "type": ["number", "null"],
                    "description": "Medida exacta pedida, si corresponde.",
                },
                "exclude_sku": {
                    "type": ["string", "null"],
                    "description": "SKU actual que no debe aparecer como alternativa.",
                },
                "available_only": {
                    "type": "boolean",
                    "description": "True para devolver solo inventario con stock y publicacion ML activa con link.",
                },
                "limit": {"type": "integer", "minimum": 1, "maximum": 20},
            },
            "required": [
                "product_terms", "context_terms", "minimum_size_mm", "exact_size_mm",
                "exclude_sku", "available_only", "limit",
            ],
            "additionalProperties": False,
        },
    },
}

_XELI_VERIFY_TOOL = {
    "type": "function",
    "function": {
        "name": "verify_ml_publication",
        "description": (
            "Verifica en vivo una publicacion ML encontrada por search_xeleria_inventory. "
            "Es obligatorio usarla antes de recomendar otro SKU o incluir su link."
        ),
        "strict": True,
        "parameters": {
            "type": "object",
            "properties": {
                "sku": {"type": "string"},
                "item_id": {"type": "string"},
            },
            "required": ["sku", "item_id"],
            "additionalProperties": False,
        },
    },
}

_KNOWLEDGE_DB_PAGE_SIZE = 1000
_KNOWLEDGE_WRITE_BATCH_SIZE = 500
_KNOWLEDGE_MAX_UPLOAD_BYTES = 15 * 1024 * 1024

_BUSINESS_WORKBOOK_SOURCE = "xeli_business_workbook"
_BUSINESS_SKILL_DEFAULTS = ["PC", "SM", "BN", "VA"]
_BUSINESS_RULE_HEADERS = [
    "SKU", "Skill", "CP Origen", "Envio", "Retiro", "Horario",
    "Es local a la calle", "Aclaración", "Modo respuesta",
]
_BUSINESS_REQUIRED_RULE_HEADERS = [
    "SKU", "Skill", "Envio", "Retiro", "Horario",
    "Es local a la calle", "Aclaración", "Modo respuesta",
]
_BUSINESS_RULE_FIELDS = [
    ("CP Origen", "código postal de origen"),
    ("Envio", "envío"),
    ("Retiro", "retiro"),
    ("Horario", "horario"),
    ("Es local a la calle", "tipo de local"),
    ("Aclaración", "aclaración"),
    ("Modo respuesta", "modo de respuesta"),
]
_BUSINESS_HELP_ROWS = [
    ("Skill ID General", "Conocimiento común. Xeli siempre lo lee antes del skill asignado al SKU."),
    ("CP Origen", "Código postal desde el que se despacha ese SKU. Si queda vacío usa el CP predeterminado."),
    ("Modo respuesta", "manual / pre_respuesta / auto_aprobada"),
    ("Mercado Envíos", "Si aplica, responder que Mercado Libre calcula costo y fecha en la publicación."),
    ("Correo / Encomienda propia", "Si aplica, responder que el costo se confirma al confeccionar etiqueta o coordinar despacho."),
    ("Stock fuente", "Xeleria o ML. Para preguntas de ML normalmente conviene ML."),
    ("Mostrar stock exacto", "No, la IA dice que hay disponibilidad sin decir cantidad."),
]


class XeliSettingsIn(BaseModel):
    greeting_enabled: bool = True
    greetings: Dict[str, str] = Field(default_factory=dict)
    signature_enabled: bool = True
    signature_text: str = ""
    quick_replies: List[Dict[str, str]] = Field(default_factory=list)


class XeliAnswerIn(BaseModel):
    text: str = Field(..., min_length=1, max_length=2000)
    greeting_enabled: Optional[bool] = None
    signature_enabled: Optional[bool] = None


class XeliResponseModeIn(BaseModel):
    response_mode: str = Field(..., min_length=1, max_length=40)


class XeliPreanswerIn(BaseModel):
    question_id: Optional[str] = None
    question_text: Optional[str] = None
    item_id: Optional[str] = None
    sku: Optional[str] = None
    question: Optional[Dict[str, Any]] = None


class XeliCorrectionIn(BaseModel):
    correction: str = Field(..., min_length=1, max_length=5000)
    sku: Optional[str] = None
    question_text: Optional[str] = None
    proposed_answer: Optional[str] = None
    item_id: Optional[str] = None


class XeliOperationalTableIn(BaseModel):
    event: Dict[str, Any]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _raw_db() -> Any:
    db = _CORE.get("_sb_raw")
    if db is None:
        sb = _CORE.get("sb")
        db = getattr(sb, "raw", sb)
    if db is None:
        raise HTTPException(status_code=500, detail="Base XelerIA no configurada")
    return db


def _error_detail(exc: Exception) -> str:
    fn = _CORE.get("pc_exception_detail")
    if callable(fn):
        try:
            return str(fn(exc))
        except Exception:
            pass
    return str(exc)


def _authorize(
    request: Request,
    token: Optional[str],
    x_admin_token: Optional[str],
    x_session_token: Optional[str],
    session_token: Optional[str],
    authorization: Optional[str],
) -> str:
    fn = _CORE.get("_inventory_auth_session_or_admin")
    if not callable(fn):
        raise HTTPException(status_code=500, detail="Autenticación XelerIA no disponible")
    session = fn(request, token, x_admin_token, x_session_token, session_token, authorization)
    tenant_from_request = _CORE.get("tenant_id_from_request")
    request_tenant = tenant_from_request(request) if callable(tenant_from_request) else None
    tenant_id = str((session or {}).get("tenant_id") or request_tenant or get_effective_tenant_id() or DEFAULT_TENANT_ID).strip()
    if not tenant_id:
        raise HTTPException(status_code=401, detail="Sesión sin tenant válido")
    return tenant_id


def _auth_kwargs(
    request: Request,
    token: Optional[str],
    x_admin_token: Optional[str],
    x_session_token: Optional[str],
    session_token: Optional[str],
    authorization: Optional[str],
) -> str:
    return _authorize(request, token, x_admin_token, x_session_token, session_token, authorization)


def _ml_get(tenant_id: str, path_or_url: str, params: Optional[Dict[str, Any]] = None, timeout: int = 45) -> Any:
    url = path_or_url if str(path_or_url).startswith("http") else f"https://api.mercadolibre.com{path_or_url}"
    fn = _CORE.get("_tenant_ml_get_json")
    if not callable(fn):
        raise HTTPException(status_code=500, detail="Cliente ML tenantizado no disponible")
    return fn(tenant_id, url, params=params or {}, timeout=timeout)


def _ml_access_token(tenant_id: str) -> str:
    fn = _CORE.get("_tenant_ml_access_token")
    token = str(fn(tenant_id) if callable(fn) else "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Mercado Libre no está conectado para este comercio")
    return token


def _seller_id(tenant_id: str) -> str:
    me = _ml_get(tenant_id, "/users/me", timeout=30)
    value = str((me or {}).get("id") or "").strip()
    if not value:
        raise HTTPException(status_code=502, detail="Mercado Libre no devolvió seller_id")
    return value


def _item_sku(item: Dict[str, Any]) -> str:
    direct = str(item.get("seller_sku") or item.get("seller_custom_field") or "").strip()
    if direct:
        return direct
    for attr in item.get("attributes") or []:
        if isinstance(attr, dict) and str(attr.get("id") or "").upper() in {"SELLER_SKU", "SKU"}:
            value = str(attr.get("value_name") or attr.get("value_id") or "").strip()
            if value:
                return value
    variations = item.get("variations") or []
    if len(variations) == 1 and isinstance(variations[0], dict):
        return _item_sku(variations[0])
    return ""


def _item_summary(tenant_id: str, item_id: str) -> Dict[str, Any]:
    if not item_id:
        return {}
    try:
        item = _ml_get(tenant_id, f"/items/{item_id}", timeout=35)
        return {
            "id": item.get("id"),
            "title": item.get("title"),
            "permalink": item.get("permalink"),
            "thumbnail": item.get("thumbnail"),
            "status": item.get("status"),
            "available_quantity": item.get("available_quantity"),
            "seller_sku": _item_sku(item),
            "shipping": item.get("shipping") or {},
        }
    except Exception as exc:
        return {"id": item_id, "error": _error_detail(exc)}


def _buyer_summary(tenant_id: str, user_id: Any) -> Dict[str, Any]:
    clean_id = str(user_id or "").strip()
    if not clean_id:
        return {}
    try:
        user = _ml_get(tenant_id, f"/users/{clean_id}", timeout=18)
    except Exception:
        return {"id": clean_id}
    address = user.get("address") if isinstance(user.get("address"), dict) else {}
    city = user.get("city") if isinstance(user.get("city"), dict) else {}
    state = user.get("state") if isinstance(user.get("state"), dict) else {}
    postal_code = str(address.get("zip_code") or "").strip()
    if not postal_code:
        try:
            addresses = _ml_get(tenant_id, f"/users/{clean_id}/addresses", timeout=18)
            if isinstance(addresses, dict):
                address_rows = [addresses]
            elif isinstance(addresses, list):
                address_rows = [row for row in addresses if isinstance(row, dict)]
            else:
                address_rows = []
            preferred = next(
                (row for row in address_rows if str(row.get("zip_code") or "").strip()),
                {},
            )
            if preferred:
                address = preferred
                postal_code = str(preferred.get("zip_code") or "").strip()
                city = preferred.get("city") if isinstance(preferred.get("city"), dict) else city
                state = preferred.get("state") if isinstance(preferred.get("state"), dict) else state
        except Exception:
            pass
    return {
        "id": clean_id,
        "nickname": user.get("nickname"),
        "country_id": user.get("country_id"),
        "city": city.get("name") or address.get("city"),
        "state": state.get("name") or address.get("state"),
        "zip_code": postal_code or None,
    }


def _questions_with_buyer_context(tenant_id: str, questions: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = [dict(question) for question in questions if isinstance(question, dict)]
    user_ids = []
    for question in rows:
        from_data = question.get("from") if isinstance(question.get("from"), dict) else {}
        user_id = str(from_data.get("id") or "").strip()
        if user_id and user_id not in user_ids:
            user_ids.append(user_id)
        if len(user_ids) >= 20:
            break
    summaries: Dict[str, Dict[str, Any]] = {}
    if user_ids:
        with ThreadPoolExecutor(max_workers=min(6, len(user_ids)), thread_name_prefix="xeli-buyers") as pool:
            futures = {user_id: pool.submit(_buyer_summary, tenant_id, user_id) for user_id in user_ids}
            for user_id, future in futures.items():
                try:
                    summaries[user_id] = future.result(timeout=20)
                except Exception:
                    summaries[user_id] = {"id": user_id}
    for question in rows:
        from_data = dict(question.get("from") or {}) if isinstance(question.get("from"), dict) else {}
        user_id = str(from_data.get("id") or "").strip()
        if user_id:
            from_data.update({key: value for key, value in (summaries.get(user_id) or {}).items() if value not in [None, ""]})
            question["from"] = from_data
    return rows


def _question_row(tenant_id: str, question: Dict[str, Any], source: str) -> Dict[str, Any]:
    item = question.get("item") if isinstance(question.get("item"), dict) else {}
    item_id = str(question.get("item_id") or item.get("id") or "").strip()
    item_data = question.get("item_data") if isinstance(question.get("item_data"), dict) else {}
    if not item_data and item_id:
        item_data = _item_summary(tenant_id, item_id)
    return {
        "tenant_id": tenant_id,
        "id": str(question.get("id") or "").strip(),
        "item_id": item_id or None,
        "seller_id": str(question.get("seller_id") or "").strip() or None,
        "status": str(question.get("status") or "UNANSWERED").upper(),
        "text": str(question.get("text") or ""),
        "date_created": question.get("date_created") or _now_iso(),
        "answer": question.get("answer"),
        "from_data": question.get("from"),
        "item_data": item_data or None,
        "raw_data": question,
        "source": source,
        "is_mock": False,
        "last_seen_at": _now_iso(),
        "updated_at": _now_iso(),
    }


def _question_public(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": row.get("id"),
        "item_id": row.get("item_id"),
        "seller_id": row.get("seller_id"),
        "status": row.get("status"),
        "text": row.get("text"),
        "date_created": row.get("date_created"),
        "answer": row.get("answer"),
        "from": row.get("from_data"),
        "item": row.get("item_data") or {},
        "source": row.get("source"),
    }


def _fetch_questions_live(tenant_id: str, status: str, limit: int = 50, offset: int = 0) -> List[Dict[str, Any]]:
    seller = _seller_id(tenant_id)
    data = _ml_get(
        tenant_id,
        "/questions/search",
        params={
            "seller_id": seller,
            "status": str(status or "UNANSWERED").upper(),
            "limit": max(1, min(int(limit or 50), 50)),
            "offset": max(0, int(offset or 0)),
            "api_version": 4,
            "sort_fields": "date_created",
            "sort_types": "DESC",
        },
        timeout=50,
    )
    rows = (data or {}).get("questions") or (data or {}).get("results") or []
    return [row for row in rows if isinstance(row, dict) and row.get("id")]


def _cache_questions(tenant_id: str, questions: Iterable[Dict[str, Any]], source: str, notify: bool) -> Dict[str, Any]:
    db = _raw_db()
    enriched_questions = _questions_with_buyer_context(tenant_id, questions)
    rows = [_question_row(tenant_id, q, source) for q in enriched_questions if q.get("id")]
    if not rows:
        return {"ok": True, "seen": 0, "new": 0, "notified": 0}
    ids = [row["id"] for row in rows]
    existing = (
        db.table("ml_questions_cache")
        .select("id")
        .eq("tenant_id", tenant_id)
        .in_("id", ids)
        .execute()
        .data
        or []
    )
    existing_ids = {str(row.get("id")) for row in existing}
    db.table("ml_questions_cache").upsert(rows, on_conflict="tenant_id,id").execute()
    new_rows = [row for row in rows if row["id"] not in existing_ids and row["status"] == "UNANSWERED"]
    auto_answered = 0
    auto_errors: List[Dict[str, str]] = []
    pending_rows = []
    for row in new_rows:
        try:
            item = row.get("item_data") or {}
            sku = str(item.get("seller_sku") or "").strip()
            mode = normalize_response_mode(_sku_skill(tenant_id, sku).get("response_mode"))
            if mode != "auto_aprobada":
                pending_rows.append(row)
                continue
            preanswer = _preanswer(tenant_id, XeliPreanswerIn(question=_question_public(row)))
            can_send = (
                bool(str(preanswer.get("response") or "").strip())
                and not bool(preanswer.get("needs_human"))
                and not bool(preanswer.get("missing_information"))
            )
            if can_send:
                _answer_question(tenant_id, row["id"], str(preanswer["response"]))
                auto_answered += 1
                continue
        except Exception as exc:
            auto_errors.append({"question_id": str(row.get("id") or ""), "error": _error_detail(exc)})
        pending_rows.append(row)
    notified = 0
    if notify and pending_rows:
        notifications = []
        for row in pending_rows:
            item = row.get("item_data") or {}
            notifications.append({
                "id": str(uuid.uuid4()),
                "tenant_id": tenant_id,
                "type": "ml_question",
                "severity": "info",
                "title": "Nueva pregunta en Mercado Libre",
                "body": str(row.get("text") or "")[:500],
                "target_view": "ml-preguntas",
                "target_url": "admin_erp.html#ml-preguntas",
                "payload": {
                    "question_id": row.get("id"),
                    "item_id": row.get("item_id"),
                    "item_title": item.get("title"),
                },
                "created_at": _now_iso(),
                "updated_at": _now_iso(),
            })
        db.table("system_notifications").insert(notifications).execute()
        notified = len(notifications)
    return {
        "ok": not auto_errors,
        "seen": len(rows),
        "new": len(new_rows),
        "auto_answered": auto_answered,
        "auto_errors": auto_errors,
        "notified": notified,
    }


def poll_ml_questions_once(limit: int = 20, dry_run: bool = False, notify: bool = True) -> Dict[str, Any]:
    """Entrada que descubre automáticamente el dispatcher cada cinco minutos."""
    tenant_id = str(get_effective_tenant_id() or DEFAULT_TENANT_ID)
    try:
        questions = _fetch_questions_live(tenant_id, "UNANSWERED", limit=max(1, min(int(limit or 20), 50)))
        if dry_run:
            return {"ok": True, "dry_run": True, "tenant_id": tenant_id, "seen": len(questions)}
        result = _cache_questions(tenant_id, questions, "xeli_5m_poller", notify=bool(notify))
        return {**result, "tenant_id": tenant_id, "version": XELI_API_VERSION}
    except Exception as exc:
        return {"ok": False, "tenant_id": tenant_id, "error": _error_detail(exc), "version": XELI_API_VERSION}


def _settings(tenant_id: str) -> Dict[str, Any]:
    rows = (
        _raw_db().table("xeli_settings")
        .select("config")
        .eq("tenant_id", tenant_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    return merged_settings((rows[0] or {}).get("config") if rows else {})


def _save_settings(tenant_id: str, config: Dict[str, Any]) -> Dict[str, Any]:
    clean = merged_settings(config)
    now = _now_iso()
    _raw_db().table("xeli_settings").upsert({
        "tenant_id": tenant_id,
        "config": clean,
        "updated_at": now,
    }, on_conflict="tenant_id").execute()
    return clean


def _knowledge(tenant_id: str, sku: str = "", include_inactive: bool = False) -> List[Dict[str, Any]]:
    clean_sku = str(sku or "").strip()
    mapping = _sku_skill(tenant_id, clean_sku) if clean_sku else {}
    assigned_skill = _clean_workbook_value(mapping.get("skill_id")).upper()
    skill_ids = ["GENERAL"]
    if assigned_skill and assigned_skill != "GENERAL":
        skill_ids.append(assigned_skill)
    q = (
        _raw_db().table("xeli_skill_instructions")
        .select("*")
        .eq("tenant_id", tenant_id)
        .in_("skill_id", skill_ids)
    )
    if not include_inactive:
        q = q.eq("active", True)
    rows = q.order("skill_order").order("sort_order").order("priority").limit(5000).execute().data or []
    return sorted(
        [
            {
                **dict(row),
                "skill_code": f"skill:{_clean_workbook_value(row.get('skill_id')).upper()}",
                "skill_type": "general" if _clean_workbook_value(row.get("skill_id")).upper() == "GENERAL" else "skill",
                "sku": clean_sku or None,
            }
            for row in rows
        ],
        key=lambda row: (
            0 if row.get("skill_type") == "general" else 1,
            int(row.get("skill_order") or 0),
            int(row.get("sort_order") or 0),
            int(row.get("priority") or 100),
        ),
    )


def _sku_skill(tenant_id: str, sku: str) -> Dict[str, Any]:
    clean_sku = _clean_workbook_value(sku)
    if not clean_sku:
        return {}
    rows = (
        _raw_db().table("xeli_sku_skills")
        .select("*")
        .eq("tenant_id", tenant_id)
        .eq("sku", clean_sku)
        .limit(1)
        .execute()
        .data
        or []
    )
    return dict(rows[0]) if rows else {}


def _response_mode_rows(tenant_id: str) -> List[Dict[str, Any]]:
    inventory = (
        _raw_db().table("inventory_items")
        .select("sku,name,active")
        .eq("tenant_id", tenant_id)
        .eq("active", True)
        .order("sku")
        .limit(10000)
        .execute()
        .data
        or []
    )
    mappings = (
        _raw_db().table("xeli_sku_skills")
        .select("sku,skill_id,response_mode")
        .eq("tenant_id", tenant_id)
        .limit(10000)
        .execute()
        .data
        or []
    )
    by_sku = {str(row.get("sku") or "").strip(): row for row in mappings if str(row.get("sku") or "").strip()}
    rows = []
    for item in inventory:
        sku = str(item.get("sku") or "").strip()
        if not sku:
            continue
        mapping = by_sku.get(sku) or {}
        rows.append({
            "sku": sku,
            "name": item.get("name"),
            "skill_id": mapping.get("skill_id"),
            "response_mode": normalize_response_mode(mapping.get("response_mode")),
        })
    return rows


def _save_response_mode(tenant_id: str, sku: str, response_mode: str) -> Dict[str, Any]:
    clean_sku = str(sku or "").strip()
    if not clean_sku:
        raise HTTPException(status_code=400, detail="Falta el SKU")
    raw_mode = re.sub(r"[\s-]+", "_", str(response_mode or "").strip().lower())
    if raw_mode not in {"manual", "pre_respuesta", "prerespuesta", "auto_aprobada", "autoaprobada"}:
        raise HTTPException(status_code=400, detail="Modo invalido. Use manual, pre_respuesta o auto_aprobada")
    clean_mode = normalize_response_mode(raw_mode)
    inventory = (
        _raw_db().table("inventory_items")
        .select("sku,name")
        .eq("tenant_id", tenant_id)
        .eq("sku", clean_sku)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not inventory:
        raise HTTPException(status_code=404, detail=f"El SKU {clean_sku} no existe en el inventario de este comercio")
    now = _now_iso()
    current = _sku_skill(tenant_id, clean_sku)
    if current:
        _raw_db().table("xeli_sku_skills").update({
            "response_mode": clean_mode,
            "updated_at": now,
        }).eq("tenant_id", tenant_id).eq("sku", clean_sku).execute()
        skill_id = current.get("skill_id")
    else:
        known = (
            _raw_db().table("xeli_skill_instructions")
            .select("skill_id")
            .eq("tenant_id", tenant_id)
            .eq("active", True)
            .limit(5000)
            .execute()
            .data
            or []
        )
        skill_id = _business_skill_from_sku(clean_sku, [row.get("skill_id") for row in known])
        _raw_db().table("xeli_sku_skills").insert({
            "tenant_id": tenant_id,
            "sku": clean_sku,
            "skill_id": skill_id,
            "response_mode": clean_mode,
            "sort_order": 0,
            "source": "xeli_training_ui",
            "source_ref": {"source": "training_ui"},
            "created_at": now,
            "updated_at": now,
        }).execute()
    return {
        "sku": clean_sku,
        "name": inventory[0].get("name"),
        "skill_id": skill_id,
        "response_mode": clean_mode,
    }


def _question_from_cache(tenant_id: str, question_id: str) -> Dict[str, Any]:
    rows = (
        _raw_db().table("ml_questions_cache")
        .select("*")
        .eq("tenant_id", tenant_id)
        .eq("id", str(question_id))
        .limit(1)
        .execute()
        .data
        or []
    )
    return _question_public(rows[0]) if rows else {}


def _live_item_context(tenant_id: str, item_id: str) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    if not item_id:
        return {}, publication_live_context({})
    item: Dict[str, Any] = {}
    description = ""
    # Título/datos vivos y descripción son lecturas independientes. Hacerlas en
    # paralelo evita sumar ambos timeouts antes de consultar a OpenAI.
    try:
        with ThreadPoolExecutor(max_workers=2, thread_name_prefix="xeli-ml") as pool:
            item_future = pool.submit(_ml_get, tenant_id, f"/items/{item_id}", None, 22)
            desc_future = pool.submit(_ml_get, tenant_id, f"/items/{item_id}/description", None, 18)
            item = item_future.result(timeout=25)
            try:
                desc = desc_future.result(timeout=22)
                description = str((desc or {}).get("plain_text") or "")[:12000]
            except Exception:
                description = ""
    except FuturesTimeoutError:
        raise HTTPException(status_code=504, detail="Mercado Libre demoró demasiado al leer la publicación")
    return item, publication_live_context(item, description)


def _inventory_context(tenant_id: str, sku: str) -> Dict[str, Any]:
    if not sku:
        return {}
    rows = (
        _raw_db().table("inventory_items")
        .select("sku,name,stock,active,item_type,variant_name,category")
        .eq("tenant_id", tenant_id)
        .eq("sku", sku)
        .limit(1)
        .execute()
        .data
        or []
    )
    item = dict(rows[0]) if rows else {}
    listings = (
        _raw_db().table("marketplace_listings")
        .select("external_product_id,external_variant_id,sku,title,price,stock,available_quantity,status,url,permalink")
        .eq("tenant_id", tenant_id)
        .eq("marketplace", "ML")
        .eq("sku", sku)
        .limit(100)
        .execute()
        .data
        or []
    )
    return {"item": item, "ml_publications": listings}


_XELI_SHIPPING_RE = re.compile(
    r"\b(env[ií]o|enviar|mandar|correo|flete|despacho|sucursal)\b",
    re.IGNORECASE,
)
_XELI_POSTAL_RE = re.compile(
    r"\b(?:cp|c[oó]digo\s+postal)?\s*:?\s*([A-Z]?\d{4}[A-Z]{0,3})\b",
    re.IGNORECASE,
)


def _question_postal_code(question_text: str) -> str:
    if not _XELI_SHIPPING_RE.search(str(question_text or "")):
        return ""
    matches = _XELI_POSTAL_RE.findall(str(question_text or ""))
    return str(matches[-1] if matches else "").strip().upper()


def _postal_location(tenant_id: str, postal_code: str) -> Dict[str, Any]:
    if not postal_code:
        return {}
    try:
        data = _ml_get(
            tenant_id,
            f"/countries/AR/zip_codes/{postal_code}",
            timeout=12,
        )
    except Exception:
        return {}
    city = data.get("city") if isinstance(data.get("city"), dict) else {}
    state = data.get("state") if isinstance(data.get("state"), dict) else {}
    extended = data.get("extended_attributes") if isinstance(data.get("extended_attributes"), dict) else {}
    return {
        "city": city.get("name") or extended.get("city_name"),
        "state": state.get("name"),
    }


def _question_destination(
    tenant_id: str,
    question: Mapping[str, Any],
    question_text: str,
) -> Dict[str, Any]:
    from_data = question.get("from") if isinstance(question.get("from"), dict) else {}
    user_id = str(from_data.get("id") or "").strip()
    if user_id and not str(from_data.get("zip_code") or "").strip():
        from_data = {**from_data, **_buyer_summary(tenant_id, user_id)}

    asked_postal_code = _question_postal_code(question_text)
    profile_postal_code = str(from_data.get("zip_code") or "").strip().upper()
    postal_code = asked_postal_code or profile_postal_code
    if not postal_code:
        return {
            "status": "needs_postal_code",
            "source": "missing",
            "city": from_data.get("city"),
            "state": from_data.get("state"),
        }

    location = _postal_location(tenant_id, postal_code)
    profile_matches = bool(profile_postal_code and profile_postal_code == postal_code)
    return {
        "status": "resolved",
        "source": "question" if asked_postal_code else "mercadolibre_buyer",
        "postal_code": postal_code,
        "city": location.get("city") or (from_data.get("city") if profile_matches else None),
        "state": location.get("state") or (from_data.get("state") if profile_matches else None),
    }


def _xeli_correo_context(
    tenant_id: str,
    question: Mapping[str, Any],
    question_text: str,
    item: Mapping[str, Any],
    origin_postal_code: Any = "",
) -> Dict[str, Any]:
    if not _XELI_SHIPPING_RE.search(str(question_text or "")):
        return {}
    destination = _question_destination(tenant_id, question, question_text)
    if destination.get("status") != "resolved":
        return {
            "source": "correo_argentino",
            "status": "needs_postal_code",
            "destination": destination,
            "instruction": (
                "No hay repregunta en Mercado Libre. Indica en esta unica respuesta que para "
                "informar el importe exacto debe enviar el codigo postal en una nueva pregunta."
            ),
        }
    client_factory = _CORE.get("_tenant_correo_client")
    client = client_factory(tenant_id) if callable(client_factory) else None
    if not client:
        return {
            "source": "correo_argentino",
            "status": "unavailable",
            "destination": destination,
        }
    dimensions = shipping_dimensions_from_ml_item(item)
    if not dimensions:
        return {
            "source": "correo_argentino",
            "status": "missing_product_dimensions",
            "destination": destination,
        }
    settings_row_fn = _CORE.get("_tenant_settings_row")
    settings_blob_fn = _CORE.get("_settings_blob")
    row = settings_row_fn(tenant_id) if callable(settings_row_fn) else {}
    settings = settings_blob_fn(row) if callable(settings_blob_fn) else {}
    try:
        quote = client.quote(
            str(destination.get("postal_code") or ""),
            dimensions,
            markup_type=settings.get("shipping_markup_type"),
            markup_value=settings.get("shipping_markup_value"),
            origin_postal_code=origin_postal_code,
        )
    except (XeleriaCorreoError, ValueError) as exc:
        return {
            "source": "correo_argentino",
            "status": "error",
            "destination": destination,
            "message": str(exc),
        }
    return {
        **quote_for_model(quote),
        "status": "quoted",
        "destination": destination,
        "dimensions": dimensions,
        "instruction": (
            "Al informar un valor incluye ciudad, provincia y codigo postal del destino usado."
        ),
    }


def _shipping_rule_key(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "").strip())
    ascii_value = "".join(char for char in normalized if not unicodedata.combining(char))
    return " ".join(ascii_value.casefold().split())


def _xeli_shipping_policy(
    item: Mapping[str, Any],
    shipping_rule: Any,
) -> Dict[str, Any]:
    shipping = item.get("shipping") if isinstance(item.get("shipping"), Mapping) else {}
    mode = str(shipping.get("mode") or "").strip().casefold()
    logistic_type = str(shipping.get("logistic_type") or "").strip()
    rule = str(shipping_rule or "").strip()
    rule_key = _shipping_rule_key(rule)

    if mode in {"me1", "me2"}:
        return {
            "kind": "mercado_envios",
            "publication_shipping_mode": mode,
            "publication_logistic_type": logistic_type or None,
            "sku_shipping_rule": rule or None,
            "instruction": (
                "Mercado Libre calcula el costo del envio. No uses ni solicites una "
                "cotizacion de Correo Argentino."
            ),
        }
    if rule_key == "envio por correo argentino directo":
        return {
            "kind": "correo_argentino_directo",
            "publication_shipping_mode": mode or None,
            "publication_logistic_type": logistic_type or None,
            "sku_shipping_rule": rule,
            "instruction": "Cotiza con Correo Argentino solamente si la pregunta consulta el envio.",
        }
    if rule_key == "requiere logistica":
        return {
            "kind": "requiere_logistica",
            "publication_shipping_mode": mode or None,
            "publication_logistic_type": logistic_type or None,
            "sku_shipping_rule": rule,
            "instruction": (
                "No cotices Correo Argentino. Indica que por el tamano del articulo la "
                "logistica se coordina despues de la compra mediante flete o empresa de transporte."
            ),
        }
    return {
        "kind": "sin_configurar",
        "publication_shipping_mode": mode or None,
        "publication_logistic_type": logistic_type or None,
        "sku_shipping_rule": rule or None,
        "instruction": "No inventes una modalidad ni un valor de envio.",
    }


def _xeli_shipping_contexts(
    tenant_id: str,
    question: Mapping[str, Any],
    question_text: str,
    item: Mapping[str, Any],
    shipping_rule: Any,
    origin_postal_code: Any = "",
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    policy = _xeli_shipping_policy(item, shipping_rule)
    correo_shipping: Dict[str, Any] = {}
    if policy["kind"] == "correo_argentino_directo":
        correo_shipping = _xeli_correo_context(
            tenant_id,
            question,
            question_text,
            item,
            origin_postal_code,
        )
    return policy, correo_shipping


def _xeli_model_options(temperature: float = 0.1) -> Tuple[str, Dict[str, Any]]:
    model = str(
        _CORE.get("XELI_OPENAI_MODEL")
        or _CORE.get("OPENAI_MODEL")
        or "gpt-4.1-mini"
    ).strip()
    effort = str(_CORE.get("XELI_REASONING_EFFORT") or "").strip().lower()
    if effort not in {"", "none", "minimal", "low", "medium", "high", "xhigh"}:
        raise HTTPException(status_code=500, detail="XELI_REASONING_EFFORT invalido")
    if model.lower().startswith("gpt-5"):
        return model, {"reasoning_effort": effort or "medium"}
    return model, {"temperature": max(0.0, min(float(temperature), 1.0))}


def _openai_json(
    system: str,
    payload: Dict[str, Any],
    temperature: float = 0.1,
    timeout_seconds: float = 20.0,
) -> Dict[str, Any]:
    api_key = str(_CORE.get("OPENAI_API_KEY") or "").strip()
    model, generation_options = _xeli_model_options(temperature)
    if not api_key:
        raise HTTPException(status_code=503, detail="OPENAI_API_KEY no configurada para Xeli")
    try:
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": model,
                "response_format": {"type": "json_object"},
                "messages": [
                    {"role": "system", "content": system},
                    {
                        "role": "user",
                        "content": json_message_payload(payload),
                    },
                ],
                **generation_options,
            },
            timeout=(10, max(1.0, min(float(timeout_seconds), 30.0))),
        )
    except requests.Timeout:
        raise HTTPException(status_code=504, detail="OpenAI demoró demasiado al preparar la respuesta de Xeli")
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"No pude comunicarme con OpenAI: {exc}")
    try:
        data = response.json()
    except Exception:
        data = {"raw": response.text[:4000]}
    if response.status_code >= 300:
        raise HTTPException(status_code=502, detail={"message": "OpenAI no respondió correctamente", "response": data})
    content = (((data.get("choices") or [{}])[0].get("message") or {}).get("content") or "{}").strip()
    try:
        parsed = json.loads(content)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"OpenAI devolvió JSON inválido: {exc}")
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=502, detail="OpenAI no devolvió un objeto JSON")
    parsed["_model"] = model
    return parsed


def _catalog_sql_terms(arguments: Dict[str, Any]) -> List[str]:
    raw_terms = arguments.get("product_terms") or arguments.get("context_terms") or []
    terms: List[str] = []
    known = set()
    for raw_value in raw_terms:
        clean = re.sub(r"[^\w\s-]+", " ", str(raw_value or ""), flags=re.UNICODE)
        clean = " ".join(clean.split()).strip()[:80]
        ascii_clean = unicodedata.normalize("NFKD", clean).encode("ascii", "ignore").decode("ascii")
        for value in [clean, ascii_clean]:
            key = value.casefold()
            if value and key not in known:
                terms.append(value)
                known.add(key)
            if len(terms) >= _XELI_SEARCH_TERM_LIMIT:
                return terms
    return terms


def _catalog_or_expression(columns: Iterable[str], terms: Iterable[str]) -> str:
    return ",".join(
        f"{column}.ilike.%{term}%"
        for term in terms
        for column in columns
    )


def _catalog_candidate_rows(
    tenant_id: str,
    arguments: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], bool]:
    terms = _catalog_sql_terms(arguments)
    if not terms:
        return [], [], False

    available_only = bool(arguments.get("available_only", True))
    inventory_query = (
        _raw_db().table("inventory_items")
        .select("sku,name,variant_name,category,stock,active,item_type")
        .eq("tenant_id", tenant_id)
        .eq("active", True)
    )
    if available_only:
        inventory_query = inventory_query.gt("stock", 0)
    inventory_direct = (
        inventory_query
        .or_(_catalog_or_expression(["sku", "name", "variant_name", "category", "item_type"], terms))
        .order("sku")
        .limit(_XELI_CANDIDATE_LIMIT + 1)
        .execute()
        .data
        or []
    )

    listing_seed = (
        _raw_db().table("marketplace_listings")
        .select("sku,title")
        .eq("tenant_id", tenant_id)
        .eq("marketplace", "ML")
        .or_(_catalog_or_expression(["sku", "title"], terms))
        .order("sku")
        .limit(_XELI_CANDIDATE_LIMIT + 1)
        .execute()
        .data
        or []
    )
    complete = len(inventory_direct) <= _XELI_CANDIDATE_LIMIT and len(listing_seed) <= _XELI_CANDIDATE_LIMIT
    inventory_direct = inventory_direct[:_XELI_CANDIDATE_LIMIT]
    listing_seed = listing_seed[:_XELI_CANDIDATE_LIMIT]

    inventory_by_sku = {
        str(row.get("sku") or "").strip().casefold(): dict(row)
        for row in inventory_direct
        if str(row.get("sku") or "").strip()
    }
    listing_skus = sorted({
        str(row.get("sku") or "").strip()
        for row in listing_seed
        if str(row.get("sku") or "").strip()
    })
    missing_skus = [sku for sku in listing_skus if sku.casefold() not in inventory_by_sku]
    if missing_skus:
        linked_query = (
            _raw_db().table("inventory_items")
            .select("sku,name,variant_name,category,stock,active,item_type")
            .eq("tenant_id", tenant_id)
            .eq("active", True)
            .in_("sku", missing_skus)
        )
        if available_only:
            linked_query = linked_query.gt("stock", 0)
        linked_rows = linked_query.limit(_XELI_CANDIDATE_LIMIT + 1).execute().data or []
        if len(linked_rows) > _XELI_CANDIDATE_LIMIT:
            complete = False
        for row in linked_rows[:_XELI_CANDIDATE_LIMIT]:
            sku_key = str(row.get("sku") or "").strip().casefold()
            if sku_key:
                inventory_by_sku.setdefault(sku_key, dict(row))

    inventory_rows = list(inventory_by_sku.values())
    if len(inventory_rows) > _XELI_CANDIDATE_LIMIT:
        complete = False
        inventory_rows = inventory_rows[:_XELI_CANDIDATE_LIMIT]
    candidate_skus = [
        str(row.get("sku") or "").strip()
        for row in inventory_rows
        if str(row.get("sku") or "").strip()
    ]
    if not candidate_skus:
        return [], [], complete

    listing_rows = (
        _raw_db().table("marketplace_listings")
        .select(
            "external_product_id,external_full_id,external_variant_id,sku,title,price,"
            "stock,available_quantity,status,url,permalink"
        )
        .eq("tenant_id", tenant_id)
        .eq("marketplace", "ML")
        .in_("sku", candidate_skus)
        .order("sku")
        .limit(_XELI_CANDIDATE_LISTING_LIMIT + 1)
        .execute()
        .data
        or []
    )
    if len(listing_rows) > _XELI_CANDIDATE_LISTING_LIMIT:
        complete = False
    return inventory_rows, listing_rows[:_XELI_CANDIDATE_LISTING_LIMIT], complete


def _catalog_search_tool(
    tenant_id: str,
    arguments: Dict[str, Any],
    catalog_cache: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    cache = catalog_cache if isinstance(catalog_cache, dict) else {}
    cache_key = json.dumps({
        "terms": _catalog_sql_terms(arguments),
        "available_only": bool(arguments.get("available_only", True)),
    }, ensure_ascii=True, sort_keys=True)
    candidate_cache = cache.setdefault("candidate_rows", {})
    cached = candidate_cache.get(cache_key)
    if cached is None:
        cached = _catalog_candidate_rows(tenant_id, arguments)
        candidate_cache[cache_key] = cached
    inventory_rows, listing_rows, catalog_complete = cached
    matches = search_inventory_catalog(
        inventory_rows,
        listing_rows,
        product_terms=arguments.get("product_terms") or [],
        context_terms=arguments.get("context_terms") or [],
        minimum_size_mm=arguments.get("minimum_size_mm"),
        exact_size_mm=arguments.get("exact_size_mm"),
        exclude_sku=arguments.get("exclude_sku") or "",
        available_only=bool(arguments.get("available_only", True)),
        limit=arguments.get("limit") or 12,
    )

    candidate_skus = [str(row.get("sku") or "").strip() for row in matches if str(row.get("sku") or "").strip()]
    mappings: List[Dict[str, Any]] = []
    if candidate_skus:
        mappings = (
            _raw_db().table("xeli_sku_skills")
            .select("sku,skill_id,response_mode")
            .eq("tenant_id", tenant_id)
            .in_("sku", candidate_skus)
            .limit(len(candidate_skus))
            .execute()
            .data
            or []
        )
    mapping_by_sku = {str(row.get("sku") or "").strip(): row for row in mappings}
    skill_ids = sorted({str(row.get("skill_id") or "").strip().upper() for row in mappings if str(row.get("skill_id") or "").strip()})
    instructions: List[Dict[str, Any]] = []
    if skill_ids:
        instructions = (
            _raw_db().table("xeli_skill_instructions")
            .select("skill_id,topic,instruction,priority,sort_order")
            .eq("tenant_id", tenant_id)
            .eq("active", True)
            .in_("skill_id", skill_ids)
            .order("skill_order")
            .order("sort_order")
            .limit(1000)
            .execute()
            .data
            or []
        )
    instructions_by_skill: Dict[str, List[Dict[str, Any]]] = {}
    for row in instructions:
        instructions_by_skill.setdefault(str(row.get("skill_id") or "").strip().upper(), []).append(dict(row))
    for row in matches:
        mapping = mapping_by_sku.get(str(row.get("sku") or "").strip()) or {}
        skill_id = str(mapping.get("skill_id") or "").strip().upper()
        row["skill_id"] = skill_id or None
        row["skill_instructions"] = instructions_by_skill.get(skill_id, [])

    return {
        "ok": True,
        "catalog_complete": catalog_complete,
        "inventory_rows_scanned": len(inventory_rows),
        "ml_listings_scanned": len(listing_rows),
        "search": {
            "product_terms": arguments.get("product_terms") or [],
            "context_terms": arguments.get("context_terms") or [],
            "minimum_size_mm": arguments.get("minimum_size_mm"),
            "exact_size_mm": arguments.get("exact_size_mm"),
            "exclude_sku": arguments.get("exclude_sku"),
            "available_only": bool(arguments.get("available_only", True)),
        },
        "items": matches,
        "total": len(matches),
    }


def _verify_ml_publication_tool(
    tenant_id: str,
    arguments: Dict[str, Any],
    allowed_candidates: set,
) -> Dict[str, Any]:
    sku = str(arguments.get("sku") or "").strip()
    item_id = str(arguments.get("item_id") or "").strip().upper()
    key = (sku.casefold(), item_id)
    if not sku or not item_id or key not in allowed_candidates:
        return {
            "ok": False,
            "available": False,
            "error": "La publicacion no proviene de la busqueda tenantizada de esta pregunta.",
            "sku": sku,
            "item_id": item_id,
        }
    rows = (
        _raw_db().table("marketplace_listings")
        .select(
            "external_product_id,external_variant_id,sku,title,price,stock,available_quantity,"
            "status,url,permalink"
        )
        .eq("tenant_id", tenant_id)
        .eq("marketplace", "ML")
        .eq("sku", sku)
        .eq("external_product_id", item_id)
        .limit(20)
        .execute()
        .data
        or []
    )
    if not rows:
        return {"ok": False, "available": False, "error": "La publicacion ya no esta vinculada al SKU.", "sku": sku, "item_id": item_id}

    live = _ml_get(tenant_id, f"/items/{item_id}", timeout=8)
    row = dict(rows[0])
    variant_id = str(row.get("external_variant_id") or "0").strip() or "0"
    live_variant = None
    if variant_id not in {"", "0"}:
        live_variant = next(
            (
                value for value in (live.get("variations") or [])
                if str((value or {}).get("id") or "").strip() == variant_id
            ),
            None,
        )
    quantity = (live_variant or {}).get("available_quantity")
    if quantity is None:
        quantity = live.get("available_quantity")
    try:
        quantity_number = float(quantity) if quantity is not None else None
    except (TypeError, ValueError):
        quantity_number = None
    status = str(live.get("status") or row.get("status") or "").strip().lower()
    permalink = str(live.get("permalink") or row.get("permalink") or row.get("url") or "").strip()
    available = status == "active" and bool(permalink) and quantity_number is not None and quantity_number > 0
    return {
        "ok": True,
        "available": available,
        "sku": sku,
        "item_id": item_id,
        "variant_id": variant_id,
        "title": row.get("title") or live.get("title"),
        "status": status,
        "available_quantity": quantity,
        "price": (live_variant or {}).get("price") if live_variant else live.get("price") or row.get("price"),
        "permalink": permalink,
    }


def _openai_preanswer_with_inventory_tools(
    tenant_id: str,
    system: str,
    payload: Dict[str, Any],
    timeout_seconds: float = 30.0,
) -> Dict[str, Any]:
    api_key = str(_CORE.get("OPENAI_API_KEY") or "").strip()
    model, generation_options = _xeli_model_options(0.1)
    if not api_key:
        raise HTTPException(status_code=503, detail="OPENAI_API_KEY no configurada para Xeli")

    messages: List[Dict[str, Any]] = [
        {"role": "system", "content": system},
        {"role": "user", "content": json_message_payload(payload)},
    ]
    deadline = time.monotonic() + max(1.0, min(float(timeout_seconds), 30.0))
    allowed_candidates: set = set()
    catalog_cache: Dict[str, Any] = {}
    searches: List[Dict[str, Any]] = []
    verified_products: List[Dict[str, Any]] = []

    for round_index in range(_XELI_TOOL_MAX_ROUNDS):
        remaining = deadline - time.monotonic()
        if remaining <= 1:
            raise HTTPException(status_code=504, detail="OpenAI demoro demasiado al consultar el inventario para Xeli")
        active_tools = [_XELI_SEARCH_TOOL] if round_index == 0 else [_XELI_SEARCH_TOOL, _XELI_VERIFY_TOOL]
        request_body: Dict[str, Any] = {
            "model": model,
            "messages": messages,
            "tools": active_tools,
            "tool_choice": "required" if round_index == 0 else "auto",
            **generation_options,
        }
        if round_index > 0:
            request_body["response_format"] = {"type": "json_object"}
        try:
            response = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json=request_body,
                timeout=(min(5.0, remaining), max(1.0, remaining)),
            )
        except requests.Timeout:
            raise HTTPException(status_code=504, detail="OpenAI demoro demasiado al consultar el inventario para Xeli")
        except requests.RequestException as exc:
            raise HTTPException(status_code=502, detail=f"No pude comunicarme con OpenAI: {exc}")
        try:
            data = response.json()
        except Exception:
            data = {"raw": response.text[:4000]}
        if response.status_code >= 300:
            raise HTTPException(status_code=502, detail={"message": "OpenAI no pudo usar las herramientas de Xeli", "response": data})

        message = (((data.get("choices") or [{}])[0]).get("message") or {})
        tool_calls = message.get("tool_calls") or []
        if not tool_calls:
            content = str(message.get("content") or "{}").strip()
            try:
                parsed = json.loads(content)
            except Exception as exc:
                raise HTTPException(status_code=502, detail=f"OpenAI devolvio JSON invalido despues de consultar inventario: {exc}")
            if not isinstance(parsed, dict):
                raise HTTPException(status_code=502, detail="OpenAI no devolvio un objeto JSON para Xeli")
            if not searches:
                raise HTTPException(status_code=502, detail="Xeli intento responder sin consultar el inventario")
            parsed["_model"] = model
            parsed["inventory_consulted"] = True
            parsed["inventory_searches"] = searches
            parsed["verified_products"] = verified_products
            parsed = validate_catalog_recommendations(parsed, verified_products)
            if any(not bool(search.get("catalog_complete")) for search in searches):
                parsed["needs_human"] = True
                missing = list(parsed.get("missing_information") or [])
                message_text = "La busqueda alcanzo el limite del catalogo y no pudo confirmar el inventario completo"
                if message_text not in missing:
                    missing.append(message_text)
                parsed["missing_information"] = missing
            return parsed

        messages.append({
            "role": "assistant",
            "content": message.get("content"),
            "tool_calls": tool_calls,
        })
        for call in tool_calls:
            call_id = str(call.get("id") or "").strip()
            fn = call.get("function") or {}
            name = str(fn.get("name") or "").strip()
            try:
                arguments = json.loads(str(fn.get("arguments") or "{}"))
                if not isinstance(arguments, dict):
                    raise ValueError("argumentos no son un objeto")
            except Exception as exc:
                tool_result = {"ok": False, "error": f"Argumentos invalidos: {exc}"}
            else:
                if name == "search_xeleria_inventory":
                    tool_result = _catalog_search_tool(tenant_id, arguments, catalog_cache)
                    searches.append({
                        "search": tool_result.get("search"),
                        "total": tool_result.get("total"),
                        "catalog_complete": tool_result.get("catalog_complete"),
                    })
                    for candidate in tool_result.get("items") or []:
                        sku_key = str(candidate.get("sku") or "").strip().casefold()
                        for publication in candidate.get("ml_publications") or []:
                            item_key = str(publication.get("item_id") or "").strip().upper()
                            if sku_key and item_key:
                                allowed_candidates.add((sku_key, item_key))
                elif name == "verify_ml_publication":
                    tool_result = _verify_ml_publication_tool(tenant_id, arguments, allowed_candidates)
                    if tool_result.get("ok"):
                        verified_products.append(tool_result)
                else:
                    tool_result = {"ok": False, "error": f"Herramienta Xeli desconocida: {name}"}
            messages.append({
                "role": "tool",
                "tool_call_id": call_id,
                "content": json_message_payload(tool_result),
            })

    raise HTTPException(status_code=502, detail="Xeli excedio el maximo de consultas permitidas al inventario")


def _operational_error_table(body: XeliOperationalTableIn) -> Dict[str, Any]:
    if len(json.dumps(body.event, ensure_ascii=False, default=str)) > 150_000:
        raise HTTPException(status_code=413, detail="El evento es demasiado grande para interpretar")
    event = operational_safe_json(body.event)
    result = _openai_json(
        (
            "Sos Xeli dentro de XelerIA. Lista los errores del JSON recibido en una tabla clara "
            "para una persona no tecnica. No diagnostiques, no opines, no recomiendes acciones y "
            "no inventes datos. Conserva referencias utiles como fila, SKU, canal, publicacion y "
            "mensaje. Traduce mensajes tecnicos a lenguaje humano: no copies diccionarios, hints "
            "SQL, nombres de tablas, roles ni codigos internos. Si varios errores son iguales, "
            "igualmente incluye cada referencia. Devuelve "
            "solo JSON con: title (string breve), intro (string breve y natural), columns (array de "
            "strings) y rows (array de arrays de strings). No incluyas columnas que esten vacias en "
            "todas las filas. Maximo 6 columnas y 100 filas."
        ),
        {"event": event},
        temperature=0.45,
        timeout_seconds=15,
    )
    table = normalize_operational_table(result)
    columns = table["columns"]
    rows = table["rows"]
    if not columns or not rows:
        raise HTTPException(status_code=502, detail="Xeli no pudo convertir el evento en una tabla")
    return {
        "ok": True,
        "title": table["title"],
        "intro": table["intro"],
        "columns": columns,
        "rows": rows,
        "model": table["model"],
        "version": XELI_API_VERSION,
    }


def _preanswer(tenant_id: str, body: XeliPreanswerIn) -> Dict[str, Any]:
    q = dict(body.question or {})
    question_id = str(body.question_id or q.get("id") or "").strip()
    if question_id and not q:
        q = _question_from_cache(tenant_id, question_id)
    question_text = str(body.question_text or q.get("text") or "").strip()
    item_id = str(body.item_id or q.get("item_id") or (q.get("item") or {}).get("id") or "").strip()
    item, live_context = _live_item_context(tenant_id, item_id)
    provided_sku = str(body.sku or (q.get("item") or {}).get("seller_sku") or "").strip()
    live_sku = str(_item_sku(item) or "").strip()
    if live_sku and provided_sku and live_sku.casefold() != provided_sku.casefold():
        raise HTTPException(status_code=409, detail=f"El SKU informado ({provided_sku}) no coincide con el SKU vivo de la publicación ({live_sku})")
    sku = live_sku or provided_sku
    knowledge = _knowledge(tenant_id, sku=sku)
    general = [row for row in knowledge if row.get("skill_type") == "general"]
    assigned_skill = [row for row in knowledge if row.get("skill_type") == "skill"]
    sku_config = _sku_skill(tenant_id, sku)
    skill_id = _clean_workbook_value(sku_config.get("skill_id")).upper()
    response_mode = normalize_response_mode(sku_config.get("response_mode"))
    inventory = _inventory_context(tenant_id, sku)
    shipping_policy, correo_shipping = _xeli_shipping_contexts(
        tenant_id,
        q,
        question_text,
        item,
        sku_config.get("shipping_rule"),
        sku_config.get("origin_postal_code"),
    )
    if not question_text:
        raise HTTPException(status_code=400, detail="Falta la pregunta para pre-responder")
    if response_mode == "manual":
        return {
            "ok": True,
            "question_id": question_id,
            "item_id": item_id,
            "sku": sku,
            "response": "",
            "response_mode": response_mode,
            "skill_id": skill_id or None,
            "needs_human": True,
            "missing_information": [],
            "skipped_ai": True,
            "version": XELI_API_VERSION,
        }
    result = _openai_preanswer_with_inventory_tools(
        tenant_id,
        (
            "Antes de responder debes consultar el inventario XelerIA con la herramienta disponible. "
            "La palabra disponible significa: SKU real del tenant, stock mayor a cero y publicacion ML activa con link; "
            "nunca significa redondear un numero. Para una medida decimal que exige el tamano disponible inmediato superior, "
            "usa minimum_size_mm y recomienda el menor matched_size_mm devuelto. Nunca inventes 13, 15 ni otra medida. "
            "Si una palabra puede referirse a productos distintos, busca por separado todas las alternativas indicadas por las reglas. "
            "Cuando una regla de conocimiento enumera significados posibles, la respuesta final debe nombrar esas opciones concretas "
            "y pedir que el comprador elija una, incluso si ninguna busqueda devuelve disponibilidad. No respondas disponibilidad "
            "definitiva usando una palabra que la propia regla declara ambigua. "
            "Antes de recomendar otro SKU o incluir su link, llama verify_ml_publication y usa solo available=true. "
            "Toda recomendacion debe incluir el permalink verificado. Si no hay coincidencia o hay ambiguedad, pregunta y marca needs_human=true. "
            "El JSON final tambien debe incluir requires_other_sku; recommended_products, una lista de objetos con sku, item_id y permalink; "
            "ambiguity_detected, un booleano; y clarification_options, una lista de las opciones concretas tomadas de las reglas. "
            "Mercado Libre no permite una conversacion con repreguntas: debes resolver en esta unica respuesta. "
            "Para responder sobre envios aplica shipping_policy como fuente autoritativa y usa una sola modalidad. "
            "Solo puedes usar correo_shipping cuando shipping_policy.kind sea correo_argentino_directo. "
            "Para cotizar Correo Argentino usa primero el codigo postal incluido en la pregunta; solo si falta usa el codigo "
            "postal del perfil del preguntador. Cuando informes un valor de envio incluye siempre la ciudad, provincia y codigo "
            "postal exactos del destino usado. Si no hay codigo postal, no inventes ni cotices: indica que debe enviarlo en una "
            "nueva pregunta. Usa exclusivamente el campo price final de correo_shipping y nunca recalcules ni menciones tarifa "
            "base, recargo, porcentaje o monto agregado. "
        ) +
        """Sos Xeli, pre-respondedor de preguntas de Mercado Libre. Leé primero TODAS las instrucciones del Skill ID GENERAL y después TODAS las instrucciones del Skill ID asociado al SKU. Leé también la configuración específica del SKU. Interpretá qué consulta el comprador usando únicamente los datos entregados. La publicación y sus datos vivos prevalecen para stock, preparación, envío y retiro. No inventes. Si falta información, indicalo para revisión humana. Devolvé JSON con: response (solo cuerpo, sin saludo ni firma), intent, needs_human, missing_information, applied_skill_codes y reasoning_brief.""",
        {
            "question": {"id": question_id, "text": question_text, "item_id": item_id},
            "read_order": ["GENERAL", skill_id or None],
            "general_skill": general,
            "assigned_skill": assigned_skill,
            "sku_config": sku_config,
            "publication_live": live_context,
            "inventory": inventory,
            "shipping_policy": shipping_policy,
            "correo_shipping": correo_shipping,
        },
    )
    return {
        "ok": True,
        "question_id": question_id,
        "item_id": item_id,
        "sku": sku,
        "response": str(result.get("response") or "").strip(),
        "response_mode": response_mode,
        "skill_id": skill_id or None,
        "needs_human": bool(result.get("needs_human")),
        "missing_information": result.get("missing_information") or [],
        "inventory_consulted": bool(result.get("inventory_consulted")),
        "inventory_searches": result.get("inventory_searches") or [],
        "recommended_products": result.get("recommended_products") or [],
        "catalog_validation_errors": result.get("catalog_validation_errors") or [],
        "ambiguity_detected": bool(result.get("ambiguity_detected")),
        "clarification_options": result.get("clarification_options") or [],
        "version": XELI_API_VERSION,
    }


def _answer_question(
    tenant_id: str,
    question_id: str,
    text: str,
    greeting_enabled: Optional[bool] = None,
    signature_enabled: Optional[bool] = None,
) -> Dict[str, Any]:
    body = str(text or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Falta el texto de la respuesta")
    settings = _settings(tenant_id)
    final_text = compose_answer(
        body,
        settings,
        greeting_enabled=greeting_enabled,
        signature_enabled=signature_enabled,
        at=datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")),
    )
    response = requests.post(
        "https://api.mercadolibre.com/answers",
        headers={"Authorization": f"Bearer {_ml_access_token(tenant_id)}", "Content-Type": "application/json"},
        json={"question_id": int(question_id), "text": final_text},
        timeout=45,
    )
    try:
        ml_data = response.json()
    except Exception:
        ml_data = {"raw": response.text[:3000]}
    if response.status_code not in {200, 201}:
        raise HTTPException(
            status_code=502,
            detail={"message": "ML no acepto la respuesta", "status_code": response.status_code, "response": ml_data},
        )
    _raw_db().table("ml_questions_cache").update({
        "status": "ANSWERED",
        "answer": ml_data,
        "updated_at": _now_iso(),
        "last_seen_at": _now_iso(),
    }).eq("tenant_id", tenant_id).eq("id", str(question_id)).execute()
    return {
        "ok": True,
        "question_id": str(question_id),
        "answer_text": final_text,
        "ml_response": ml_data,
        "version": XELI_API_VERSION,
    }


def _translate_correction(tenant_id: str, skill_type: str, body: XeliCorrectionIn) -> Dict[str, Any]:
    sku = str(body.sku or "").strip()
    if skill_type == "sku" and body.item_id:
        item = _ml_get(tenant_id, f"/items/{str(body.item_id).strip()}", timeout=35)
        live_sku = str(_item_sku(item) or "").strip()
        if live_sku and sku and live_sku.casefold() != sku.casefold():
            raise HTTPException(status_code=409, detail=f"El SKU informado ({sku}) no coincide con el SKU vivo de la publicación ({live_sku})")
        sku = live_sku or sku
    if skill_type == "sku" and not sku:
        raise HTTPException(status_code=400, detail="Corregir skill requiere el SKU exacto")
    skill_id = "GENERAL"
    if skill_type == "sku":
        mapping = _sku_skill(tenant_id, sku)
        skill_id = _clean_workbook_value(mapping.get("skill_id")).upper()
        if not skill_id:
            raise HTTPException(status_code=400, detail=f"El SKU {sku} no tiene un Skill ID asignado")
    existing_rows = (
        _raw_db().table("xeli_skill_instructions")
        .select("id,topic,instruction,priority,skill_order,sort_order")
        .eq("tenant_id", tenant_id)
        .eq("skill_id", skill_id)
        .eq("active", True)
        .order("sort_order")
        .limit(500)
        .execute()
        .data
        or []
    )
    result = _openai_json(
        """Convertí la corrección humana en UNA regla dura para Xeli: breve, inequívoca, imperativa y reusable. Resumí; no expliques motivos, no agregues pasos, verificaciones, excepciones ni recomendaciones que el humano no haya escrito. Si el texto ya es breve, sólo normalizalo como regla. Compará con las reglas existentes: si corrige, contradice o mejora una de ellas, devolvé action='replace' y su replace_id; si es conocimiento nuevo, action='add'. Devolvé únicamente JSON con action, replace_id, topic, instruction y priority. La instruction debe ser una sola regla y normalmente no superar 300 caracteres.""",
        {
            "target_skill": skill_id,
            "human_correction": body.correction,
            "question": body.question_text,
            "proposed_answer": body.proposed_answer,
            "item_id": body.item_id,
            "existing_rules": existing_rows,
        },
    )
    instruction = _clean_workbook_value(result.get("instruction"))
    if not instruction:
        raise HTTPException(status_code=502, detail="OpenAI no devolvió una instrucción técnica")
    valid_existing = {str(row.get("id")): row for row in existing_rows if row.get("id")}
    replace_id = _clean_workbook_value(result.get("replace_id"))
    replace_row = valid_existing.get(replace_id) if _clean_workbook_value(result.get("action")).lower() == "replace" else None
    values = {
        "topic": _clean_workbook_value(result.get("topic") or "respuesta").lower(),
        "instruction": instruction,
        "priority": int(result.get("priority") or 100),
        "active": True,
        "source": "xeli_training_correction",
        "source_ref": {
            "human_correction": body.correction,
            "question_text": body.question_text,
            "proposed_answer": body.proposed_answer,
            "item_id": body.item_id,
            "sku": sku or None,
            "openai_model": result.get("_model"),
            "replaced_instruction_id": replace_id or None,
        },
        "updated_at": _now_iso(),
    }
    if replace_row:
        _raw_db().table("xeli_skill_instructions").update(values).eq(
            "tenant_id", tenant_id
        ).eq("id", replace_id).execute()
        changed_existing = True
    else:
        skill_order = 0 if skill_id == "GENERAL" else min(
            [int(row.get("skill_order") or 999999) for row in existing_rows] or [999999]
        )
        sort_order = max([int(row.get("sort_order") or 0) for row in existing_rows] or [0]) + 1
        row = {
            "id": str(uuid.uuid4()), "tenant_id": tenant_id, "skill_id": skill_id,
            "skill_order": skill_order, "sort_order": sort_order,
            "created_at": _now_iso(), **values,
        }
        _raw_db().table("xeli_skill_instructions").insert(row).execute()
        changed_existing = False
    return {
        "ok": True,
        "message": f"Regla guardada en {skill_id}.",
        "skill_id": skill_id,
        "rule": instruction,
        "changed_existing": changed_existing,
        "version": XELI_API_VERSION,
    }


def _training_rows(tenant_id: str, limit: int = 50, refresh: bool = False) -> Tuple[List[Dict[str, Any]], List[str]]:
    warnings: List[str] = []
    if refresh:
        for status in ("UNANSWERED", "ANSWERED"):
            try:
                live = _fetch_questions_live(tenant_id, status, limit=min(50, limit))
                _cache_questions(tenant_id, live, "xeli_training_refresh", notify=status == "UNANSWERED")
            except Exception as exc:
                warnings.append(f"{status}: {_error_detail(exc)}")
    rows = (
        _raw_db().table("ml_questions_cache")
        .select("*")
        .eq("tenant_id", tenant_id)
        .order("date_created", desc=True)
        .limit(max(1, min(int(limit or 50), 200)))
        .execute()
        .data
        or []
    )
    return [_question_public(row) for row in rows], warnings


def _clean_workbook_value(value: Any) -> str:
    return str(value or "").strip()


def _workbook_header_key(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", _clean_workbook_value(value))
    return " ".join(
        "".join(char for char in normalized if not unicodedata.combining(char))
        .casefold()
        .split()
    )


def _business_skill_from_sku(sku: Any, known_skills: Iterable[str] = ()) -> str:
    clean_sku = _clean_workbook_value(sku).upper()
    candidates = sorted(
        {_clean_workbook_value(value).upper() for value in known_skills if _clean_workbook_value(value)},
        key=len,
        reverse=True,
    )
    for skill in candidates:
        if clean_sku.startswith(skill):
            return skill
    match = re.match(r"[A-Z]+", clean_sku)
    return (match.group(0)[:2] if match else "SKU") or "SKU"


def _business_workbook_from_rows(rows: Iterable[Dict[str, Any]]) -> bytes:
    """Arma el Excel humano original sin desorganizar la tabla técnica interna."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Falta openpyxl: {exc}")

    rule_rows: Dict[str, Dict[str, Any]] = {}
    skill_entries: Dict[Tuple[str, int, str], str] = {}
    skill_order: Dict[str, int] = {skill: index for index, skill in enumerate(_BUSINESS_SKILL_DEFAULTS, 1)}
    prompt_entries: Dict[Tuple[int, str], str] = {}

    def ensure_rule(sku: Any, skill: Any = "", order: Any = 999999) -> Dict[str, Any]:
        clean_sku = _clean_workbook_value(sku)
        if not clean_sku:
            return {}
        current = rule_rows.setdefault(
            clean_sku,
            {header: "" for header in _BUSINESS_RULE_HEADERS} | {"SKU": clean_sku, "_order": 999999},
        )
        clean_skill = _clean_workbook_value(skill).upper()
        if clean_skill and not current.get("Skill"):
            current["Skill"] = clean_skill
        try:
            current["_order"] = min(int(current.get("_order") or 999999), int(order or 999999))
        except (TypeError, ValueError):
            pass
        return current

    materialized = [dict(row) for row in (rows or []) if isinstance(row, dict) and row.get("active", True)]
    for row in materialized:
        ref = row.get("source_ref") if isinstance(row.get("source_ref"), dict) else {}
        kind = _clean_workbook_value(ref.get("business_kind")).lower()
        if kind in {"rule", "mapping"}:
            business_row = ref.get("business_row") if isinstance(ref.get("business_row"), dict) else {}
            sku = business_row.get("SKU") or row.get("sku")
            target = ensure_rule(sku, business_row.get("Skill") or ref.get("business_skill"), ref.get("workbook_row"))
            for header in _BUSINESS_RULE_HEADERS:
                value = business_row.get(header)
                if value not in (None, "") and target:
                    target[header] = _clean_workbook_value(value)
            continue
        if kind == "skill":
            skill = _clean_workbook_value(ref.get("business_skill")).upper()
            value = _clean_workbook_value(ref.get("business_value") or row.get("instruction"))
            try:
                index = int(ref.get("business_skill_index") or 999999)
            except (TypeError, ValueError):
                index = 999999
            if skill and value:
                skill_entries[(skill, index, value)] = value
                try:
                    skill_order[skill] = min(skill_order.get(skill, 999999), int(ref.get("business_skill_order") or 999999))
                except (TypeError, ValueError):
                    pass
            continue
        if kind == "prompt":
            value = _clean_workbook_value(ref.get("business_value") or row.get("instruction"))
            try:
                index = int(ref.get("workbook_row") or 999999)
            except (TypeError, ValueError):
                index = 999999
            if value:
                skill_entries[("GENERAL", index, value)] = value
                skill_order["GENERAL"] = 0
            continue

        instruction = _clean_workbook_value(row.get("instruction"))
        if not instruction:
            continue
        if _clean_workbook_value(row.get("skill_type")).lower() == "general":
            skill_entries[("GENERAL", 999999, instruction)] = instruction
            skill_order["GENERAL"] = 0
            continue
        sku = _clean_workbook_value(row.get("sku"))
        if not sku:
            continue
        target = ensure_rule(sku)
        existing = _clean_workbook_value(target.get("Aclaración"))
        if instruction not in existing:
            target["Aclaración"] = f"{existing}\n{instruction}".strip()

    known_skills = set(_BUSINESS_SKILL_DEFAULTS)
    known_skills.update(skill for skill, _, _ in skill_entries)
    known_skills.update(_clean_workbook_value(row.get("Skill")).upper() for row in rule_rows.values())
    for row in rule_rows.values():
        if not row.get("Skill"):
            row["Skill"] = _business_skill_from_sku(row.get("SKU"), known_skills)

    wb = Workbook()
    ws_rules = wb.active
    ws_rules.title = "Reglas SKU"
    ws_rules.append(_BUSINESS_RULE_HEADERS)
    for row in sorted(rule_rows.values(), key=lambda value: (int(value.get("_order") or 999999), value.get("SKU") or "")):
        ws_rules.append([row.get(header, "") for header in _BUSINESS_RULE_HEADERS])

    ws_skills = wb.create_sheet("Skills")
    ws_skills.append(["Skill ID", "Skill instrucción"])
    for (skill, index, value), _ in sorted(
        skill_entries.items(), key=lambda item: (skill_order.get(item[0][0], 999999), item[0][0], item[0][1], item[0][2])
    ):
        ws_skills.append([skill, value])

    ws_help = wb.create_sheet("Ayuda")
    ws_help.append(["Campo", "Uso"])
    for row in _BUSINESS_HELP_ROWS:
        ws_help.append(list(row))

    header_fill = PatternFill("solid", fgColor="DCC58F")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(name="Calibri", size=11, bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center")

    for column, width in {
        "A": 24, "B": 16, "C": 14, "D": 34, "E": 42,
        "F": 38, "G": 20, "H": 44, "I": 20,
    }.items():
        ws_rules.column_dimensions[column].width = width
    ws_skills.column_dimensions["A"].width = 18
    ws_skills.column_dimensions["B"].width = 120
    for row_number in range(2, ws_skills.max_row + 1):
        value = _clean_workbook_value(ws_skills.cell(row=row_number, column=2).value)
        ws_skills.cell(row=row_number, column=1).alignment = Alignment(vertical="top")
        ws_skills.cell(row=row_number, column=2).alignment = Alignment(vertical="top", wrap_text=True)
        ws_skills.row_dimensions[row_number].height = 15 * max(1, (len(value) + 109) // 110)
    ws_help.column_dimensions["A"].width = 28
    ws_help.column_dimensions["B"].width = 42

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _business_workbook_from_normalized(
    sku_rows: Iterable[Dict[str, Any]], instruction_rows: Iterable[Dict[str, Any]],
) -> bytes:
    workbook_rows: List[Dict[str, Any]] = []
    for row in sku_rows or []:
        business_row = {
            "SKU": row.get("sku") or "",
            "Skill": row.get("skill_id") or "",
            "CP Origen": row.get("origin_postal_code") or "",
            "Envio": row.get("shipping_rule") or "",
            "Retiro": row.get("pickup") or "",
            "Horario": row.get("schedule") or "",
            "Es local a la calle": row.get("street_store") or "",
            "Aclaración": row.get("clarification") or "",
            "Modo respuesta": row.get("response_mode") or "",
        }
        workbook_rows.append({
            "active": True,
            "sku": row.get("sku"),
            "source_ref": {
                "business_kind": "mapping", "business_skill": row.get("skill_id"),
                "business_row": business_row, "workbook_row": row.get("sort_order") or 999999,
            },
        })
    for row in instruction_rows or []:
        skill_id = _clean_workbook_value(row.get("skill_id")).upper()
        ref = row.get("source_ref") if isinstance(row.get("source_ref"), dict) else {}
        if skill_id == "GENERAL":
            source_ref = {
                **ref, "business_kind": "prompt", "business_value": row.get("instruction"),
                "workbook_row": row.get("sort_order") or 999999,
            }
        else:
            source_ref = {
                **ref, "business_kind": "skill", "business_skill": skill_id,
                "business_skill_index": row.get("sort_order") or 999999,
                "business_skill_order": row.get("skill_order") or 999999,
                "business_value": row.get("instruction"),
            }
        workbook_rows.append({
            "active": bool(row.get("active", True)),
            "instruction": row.get("instruction"),
            "source_ref": source_ref,
        })
    return _business_workbook_from_rows(workbook_rows)


def _knowledge_db_execute(build_query: Any, operation: str, attempts: int = 4) -> Any:
    """Reintenta sólo fallas transitorias y recrea la consulta en cada intento."""
    last_error: Optional[Exception] = None
    transient_markers = (
        "connectionterminated", "connection terminated", "server disconnected",
        "connection reset", "timeout", "timed out", "temporarily unavailable",
        "bad gateway", "gateway timeout", "error_code:1", "502", "503", "504",
    )
    for attempt in range(max(1, attempts)):
        try:
            return build_query().execute()
        except Exception as exc:
            last_error = exc
            detail = _error_detail(exc).lower()
            if not any(marker in detail for marker in transient_markers) or attempt + 1 >= attempts:
                break
            time.sleep(0.35 * (2 ** attempt))
    raise HTTPException(
        status_code=503,
        detail=f"No pude completar {operation} después de reintentar: {_error_detail(last_error)}",
    )


def _knowledge_table_rows(
    table: str,
    tenant_id: str,
    *,
    select_columns: str = "*",
    active_only: bool = False,
    source: Optional[str] = None,
    order_columns: Tuple[str, ...] = (),
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    offset = 0
    while True:
        start = offset

        def build_query() -> Any:
            query = _raw_db().table(table).select(select_columns).eq("tenant_id", tenant_id)
            if active_only:
                query = query.eq("active", True)
            if source:
                query = query.eq("source", source)
            for column in order_columns:
                query = query.order(column)
            return query.range(start, start + _KNOWLEDGE_DB_PAGE_SIZE - 1)

        page = _knowledge_db_execute(build_query, f"la lectura de {table}").data or []
        rows.extend(page)
        if len(page) < _KNOWLEDGE_DB_PAGE_SIZE:
            return rows
        offset += len(page)


def _knowledge_workbook(tenant_id: str) -> bytes:
    sku_rows = _knowledge_table_rows(
        "xeli_sku_skills",
        tenant_id,
        order_columns=("sort_order", "sku"),
    )
    instruction_rows = _knowledge_table_rows(
        "xeli_skill_instructions",
        tenant_id,
        active_only=True,
        order_columns=("skill_order", "skill_id", "sort_order"),
    )
    return _business_workbook_from_normalized(sku_rows, instruction_rows)


def _bool_cell(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().lower() in {"1", "true", "si", "sí", "yes", "activo", "activa"}


def _technical_knowledge_records(wb: Any, tenant_id: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    ws = wb["Conocimiento Xeli"]
    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
    required = {"skill type", "tema", "instrucción"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="Faltan columnas Skill type, Tema o Instrucción")
    index = {name: idx for idx, name in enumerate(headers)}
    records: List[Dict[str, Any]] = []
    errors: List[str] = []
    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(value not in (None, "") for value in values):
            continue
        try:
            get = lambda name: values[index[name]] if name in index and index[name] < len(values) else None
            source_ref = {}
            raw_ref = get("trazabilidad json")
            if raw_ref:
                parsed = json.loads(str(raw_ref))
                source_ref = parsed if isinstance(parsed, dict) else {}
            record = knowledge_record(
                skill_type=get("skill type"),
                sku=get("sku"),
                topic=get("tema"),
                instruction=get("instrucción"),
                priority=get("prioridad") or 100,
                active=_bool_cell(get("activo")) if get("activo") not in (None, "") else True,
                source=get("origen") or "excel_import",
                source_ref=source_ref,
            )
            record.update({"tenant_id": tenant_id, "updated_at": _now_iso()})
            row_id = str(get("id") or "").strip()
            record.update({"id": row_id or str(uuid.uuid4()), "created_at": _now_iso()})
            records.append(record)
        except Exception as exc:
            errors.append(f"Fila {row_number}: {_error_detail(exc)}")
    return records, errors


def _business_record_id(tenant_id: str, *parts: Any) -> str:
    key = "|".join(["xeli-business-workbook", tenant_id] + [_clean_workbook_value(part) for part in parts])
    return str(uuid.uuid5(uuid.NAMESPACE_URL, key))


def _business_instruction(header: str, sku: str, value: str) -> str:
    templates = {
        "CP Origen": "Para el SKU {sku}, el código postal de origen es: {value}",
        "Envio": "Para el SKU {sku}, la regla de envío es: {value}",
        "Modo respuesta": "Para el SKU {sku}, el modo de respuesta configurado es: {value}.",
        "Retiro": "Para el SKU {sku}, la regla de retiro es: {value}",
        "Horario": "Para el SKU {sku}, el horario aplicable es: {value}",
        "Es local a la calle": "Para el SKU {sku}, la condición del local es: {value}",
        "Aclaración": "Para el SKU {sku}, tener en cuenta: {value}",
    }
    return templates[header].format(sku=sku, value=value)


def _save_knowledge_records(
    tenant_id: str,
    records: List[Dict[str, Any]],
    errors: List[str],
    *,
    replace_business: bool = False,
) -> Dict[str, Any]:
    if errors:
        return {"ok": False, "inserted": 0, "updated": 0, "removed": 0, "skipped": 0, "errors": errors[:100]}
    db = _raw_db()
    replace = getattr(db, "replace_tenant_knowledge", None)
    if not callable(replace):
        raise HTTPException(status_code=500, detail="La base no admite reemplazo atomico de conocimiento")
    counts = replace(tenant_id, {"xeli_knowledge": records})
    removed = sum(item["removed"] for item in counts.values())
    return {
        "ok": True, "inserted": len(records), "updated": 0, "removed": removed,
        "skipped": 0, "errors": [], "tenant_id": tenant_id,
        "format": "xeli_business_workbook" if replace_business else "technical",
    }


def _business_normalized_rows(
    wb: Any, tenant_id: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    required_sheets = {"Reglas SKU", "Skills"}
    missing = sorted(required_sheets.difference(wb.sheetnames))
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan hojas en el Excel: {', '.join(missing)}")
    now = _now_iso()
    errors: List[str] = []
    sku_rows: List[Dict[str, Any]] = []
    instruction_rows: List[Dict[str, Any]] = []

    ws_rules = wb["Reglas SKU"]
    headers = [_clean_workbook_value(cell.value) for cell in ws_rules[1]]
    header_index = {_workbook_header_key(header): index for index, header in enumerate(headers)}
    missing_headers = [
        header for header in _BUSINESS_REQUIRED_RULE_HEADERS
        if _workbook_header_key(header) not in header_index
    ]
    if missing_headers:
        raise HTTPException(status_code=400, detail=f"Faltan columnas en 'Reglas SKU': {', '.join(missing_headers)}")
    seen_skus: set[str] = set()
    known_skills = set(_BUSINESS_SKILL_DEFAULTS)
    for row_number, values in enumerate(ws_rules.iter_rows(min_row=2, values_only=True), 2):
        row = {
            header: (
                _clean_workbook_value(values[header_index[_workbook_header_key(header)]])
                if _workbook_header_key(header) in header_index
                and header_index[_workbook_header_key(header)] < len(values)
                else ""
            )
            for header in _BUSINESS_RULE_HEADERS
        }
        if not any(row.values()):
            continue
        sku = row["SKU"]
        if not sku:
            errors.append(f"Reglas SKU, fila {row_number}: falta SKU")
            continue
        if sku in seen_skus:
            errors.append(f"Reglas SKU, fila {row_number}: SKU duplicado {sku}")
            continue
        seen_skus.add(sku)
        skill_id = row["Skill"].upper() or _business_skill_from_sku(sku, known_skills)
        known_skills.add(skill_id)
        try:
            origin_postal_code = (
                normalize_postal_code(row["CP Origen"])
                if row["CP Origen"]
                else None
            )
        except ValueError as exc:
            errors.append(f"Reglas SKU, fila {row_number}: {sku}, {exc}")
            continue
        sku_rows.append({
            "tenant_id": tenant_id,
            "sku": sku,
            "skill_id": skill_id,
            "origin_postal_code": origin_postal_code,
            "shipping_rule": row["Envio"] or None,
            "response_mode": row["Modo respuesta"] or None,
            "pickup": row["Retiro"] or None,
            "schedule": row["Horario"] or None,
            "street_store": row["Es local a la calle"] or None,
            "clarification": row["Aclaración"] or None,
            "sort_order": row_number - 1,
            "source": _BUSINESS_WORKBOOK_SOURCE,
            "source_ref": {"workbook_sheet": "Reglas SKU", "workbook_row": row_number},
            "created_at": now,
            "updated_at": now,
        })

    ws_skills = wb["Skills"]
    skill_values: List[Tuple[str, int, int, str, str]] = []
    if (
        _clean_workbook_value(ws_skills["A1"].value).lower() in {"skill", "skill id"}
        and _clean_workbook_value(ws_skills["B1"].value).lower()
        in {"instrucción", "instruccion", "skill instrucción", "skill instruccion"}
    ):
        per_skill_index: Dict[str, int] = {}
        skill_order: Dict[str, int] = {}
        for row_number in range(2, ws_skills.max_row + 1):
            skill_id = _clean_workbook_value(ws_skills.cell(row=row_number, column=1).value).upper()
            instruction = _clean_workbook_value(ws_skills.cell(row=row_number, column=2).value)
            if not skill_id and not instruction:
                continue
            if not skill_id or not instruction:
                errors.append(f"Skills, fila {row_number}: falta Skill ID o Skill instrucción")
                continue
            if skill_id not in skill_order:
                skill_order[skill_id] = 0 if skill_id == "GENERAL" else len(
                    [value for value in skill_order if value != "GENERAL"]
                ) + 1
            per_skill_index[skill_id] = per_skill_index.get(skill_id, 0) + 1
            skill_values.append((skill_id, skill_order[skill_id], per_skill_index[skill_id], instruction, f"B{row_number}"))
    else:
        # Compatibilidad con la planilla inicial: un Skill ID por columna.
        for column_index, cell in enumerate(ws_skills[1], 1):
            skill_id = _clean_workbook_value(cell.value).upper()
            if not skill_id:
                continue
            skill_index = 0
            for row_number in range(2, ws_skills.max_row + 1):
                instruction = _clean_workbook_value(ws_skills.cell(row=row_number, column=column_index).value)
                if not instruction:
                    continue
                skill_index += 1
                skill_values.append((skill_id, column_index, skill_index, instruction, f"{cell.column_letter}{row_number}"))

    for skill_id, skill_order, sort_order, instruction, workbook_cell in skill_values:
        instruction_rows.append({
            "id": _business_record_id(tenant_id, "instruction", skill_id, sort_order),
            "tenant_id": tenant_id,
            "skill_id": skill_id,
            "topic": f"skill {skill_id.lower()} {sort_order:03d}",
            "instruction": instruction,
            "priority": 100,
            "skill_order": skill_order,
            "sort_order": sort_order,
            "active": True,
            "source": _BUSINESS_WORKBOOK_SOURCE,
            "source_ref": {"workbook_sheet": "Skills", "workbook_cell": workbook_cell},
            "created_at": now,
            "updated_at": now,
        })

    # Compatibilidad con el archivo original: Prompts IA se convierte en Skill ID General.
    if "Prompts IA" in wb.sheetnames:
        ws_prompts = wb["Prompts IA"]
        if _clean_workbook_value(ws_prompts["A1"].value).upper() != "PROMPTS":
            errors.append("Prompts IA: la celda A1 debe decir PROMPTS")
        existing_general = sum(1 for row in instruction_rows if row.get("skill_id") == "GENERAL")
        prompt_index = existing_general
        for row_number in range(2, ws_prompts.max_row + 1):
            instruction = _clean_workbook_value(ws_prompts.cell(row=row_number, column=1).value)
            if not instruction:
                continue
            prompt_index += 1
            instruction_rows.append({
                "id": _business_record_id(tenant_id, "instruction", "GENERAL", prompt_index),
                "tenant_id": tenant_id,
                "skill_id": "GENERAL",
                "topic": f"skill general {prompt_index:03d}",
                "instruction": instruction,
                "priority": 100,
                "skill_order": 0,
                "sort_order": prompt_index,
                "active": True,
                "source": _BUSINESS_WORKBOOK_SOURCE,
                "source_ref": {"workbook_sheet": "Prompts IA", "workbook_row": row_number},
                "created_at": now,
                "updated_at": now,
            })
    return sku_rows, instruction_rows, errors


def _save_business_normalized(
    tenant_id: str,
    sku_rows: List[Dict[str, Any]],
    instruction_rows: List[Dict[str, Any]],
    errors: List[str],
) -> Dict[str, Any]:
    if errors:
        return {
            "ok": False, "inserted": 0, "updated": 0, "removed": 0,
            "sku_rows": 0, "skill_instructions": 0, "errors": errors[:100],
        }
    db = _raw_db()
    replace = getattr(db, "replace_tenant_knowledge", None)
    if not callable(replace):
        raise HTTPException(status_code=500, detail="La base no admite reemplazo atomico de conocimiento")
    counts = replace(
        tenant_id,
        {
            "xeli_sku_skills": sku_rows,
            "xeli_skill_instructions": instruction_rows,
        },
    )
    removed = sum(item["removed"] for item in counts.values())
    inserted = len(sku_rows) + len(instruction_rows)
    return {
        "ok": True,
        "tenant_id": tenant_id,
        "inserted": inserted,
        "updated": 0,
        "removed": removed,
        "sku_rows": len(sku_rows),
        "skill_instructions": len(instruction_rows),
        "errors": [],
        "format": "xeli_normalized_business_workbook",
    }


def _import_knowledge(tenant_id: str, content: bytes) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Falta openpyxl: {exc}")
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {exc}")
    if "Conocimiento Xeli" in wb.sheetnames:
        records, errors = _technical_knowledge_records(wb, tenant_id)
        return _save_knowledge_records(tenant_id, records, errors)
    sku_rows, instruction_rows, errors = _business_normalized_rows(wb, tenant_id)
    return _save_business_normalized(tenant_id, sku_rows, instruction_rows, errors)


def install_xeli_routes(router: Any, core: Dict[str, Any]) -> None:
    global _CORE
    _CORE = core
    core["poll_ml_questions_once"] = poll_ml_questions_once

    def auth(
        request: Request,
        token: Optional[str], x_admin_token: Optional[str], x_session_token: Optional[str],
        session_token: Optional[str], authorization: Optional[str],
    ) -> str:
        return _auth_kwargs(request, token, x_admin_token, x_session_token, session_token, authorization)

    @router.get("/admin/xeli/settings")
    def xeli_settings_get(request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        return {"ok": True, "tenant_id": tenant_id, "settings": _settings(tenant_id), "version": XELI_API_VERSION}

    @router.put("/admin/xeli/settings")
    def xeli_settings_put(body: XeliSettingsIn, request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        payload = body.model_dump() if hasattr(body, "model_dump") else body.dict()
        return {"ok": True, "tenant_id": tenant_id, "settings": _save_settings(tenant_id, payload), "version": XELI_API_VERSION}

    @router.get("/admin/xeli/ml/questions")
    def xeli_questions_get(request: Request, status: str = Query("UNANSWERED"), limit: int = Query(20, ge=1, le=50), offset: int = Query(0, ge=0), refresh: bool = Query(False), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        warning = None
        if refresh:
            try:
                live = _fetch_questions_live(tenant_id, status, limit=limit, offset=offset)
                _cache_questions(tenant_id, live, "xeli_manual_refresh", notify=str(status).upper() == "UNANSWERED")
            except Exception as exc:
                warning = _error_detail(exc)
        q = _raw_db().table("ml_questions_cache").select("*", count="exact").eq("tenant_id", tenant_id)
        if str(status).upper() != "ALL":
            q = q.eq("status", str(status).upper())
        result = q.order("date_created", desc=True).range(offset, offset + limit - 1).execute()
        rows = result.data or []
        total = result.count if result.count is not None else len(rows)
        return {"ok": True, "items": [_question_public(row) for row in rows], "total": total, "limit": limit, "offset": offset, "has_more": offset + len(rows) < int(total or 0), "warning": warning, "source": "XelerIA tenant cache", "version": XELI_API_VERSION}

    @router.post("/admin/xeli/ml/questions/{question_id}/answer")
    def xeli_question_answer(question_id: str, body: XeliAnswerIn, request: Request, dry_run: bool = Query(False), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        settings = _settings(tenant_id)
        final_text = compose_answer(body.text, settings, greeting_enabled=body.greeting_enabled, signature_enabled=body.signature_enabled, at=datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")))
        if dry_run:
            return {"ok": True, "dry_run": True, "answer_text": final_text}
        return _answer_question(
            tenant_id,
            question_id,
            body.text,
            greeting_enabled=body.greeting_enabled,
            signature_enabled=body.signature_enabled,
        )

    @router.get("/admin/xeli/response-modes")
    def xeli_response_modes_get(request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        rows = _response_mode_rows(tenant_id)
        return {"ok": True, "items": rows, "total": len(rows), "version": XELI_API_VERSION}

    @router.put("/admin/xeli/response-modes/{sku}")
    def xeli_response_mode_put(sku: str, body: XeliResponseModeIn, request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        item = _save_response_mode(tenant_id, sku, body.response_mode)
        return {"ok": True, "item": item, "version": XELI_API_VERSION}

    @router.get("/admin/xeli/ml/link-search")
    def xeli_link_search(request: Request, q: str = Query(..., min_length=4), limit: int = Query(10000, ge=1, le=10000), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        rows = _raw_db().table("marketplace_listings").select("external_product_id,external_variant_id,sku,title,url,permalink,status").eq("tenant_id", tenant_id).eq("marketplace", "ML").limit(10000).execute().data or []
        found = filter_publications_by_title(rows, q)[:limit]
        return {"ok": True, "query": q, "search_field": "title", "items": [{**row, "link": row.get("permalink") or row.get("url")} for row in found], "total": len(found), "version": XELI_API_VERSION}

    @router.get("/admin/xeli/training/questions")
    def xeli_training_questions(request: Request, limit: int = Query(50, ge=1, le=200), refresh: bool = Query(False), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        rows, warnings = _training_rows(tenant_id, limit=limit, refresh=refresh)
        return {"ok": True, "items": rows, "total": len(rows), "warnings": warnings, "version": XELI_API_VERSION}

    @router.get("/admin/xeli/training/questions.txt")
    def xeli_training_questions_txt(request: Request, limit: int = Query(50, ge=1, le=200), refresh: bool = Query(False), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        rows, warnings = _training_rows(tenant_id, limit=limit, refresh=refresh)
        chunks = []
        for idx, row in enumerate(rows, 1):
            item = row.get("item") or {}
            answer = row.get("answer") or {}
            chunks.append("\n".join([
                f"[{idx}] Pregunta {row.get('id') or ''}",
                f"Fecha: {row.get('date_created') or ''}",
                f"Item: {row.get('item_id') or ''}",
                f"SKU: {item.get('seller_sku') or ''}",
                f"Publicación: {item.get('title') or ''}",
                f"Pregunta: {row.get('text') or ''}",
                f"Respuesta: {answer.get('text') or '' if isinstance(answer, dict) else answer}",
            ]))
        if warnings:
            chunks.append("Advertencias de actualización:\n" + "\n".join(warnings))
        headers = {"Content-Disposition": 'attachment; filename="ultimas_preguntas_xeli.txt"'}
        return PlainTextResponse("\n\n".join(chunks), headers=headers)

    @router.post("/admin/xeli/training/preanswer")
    def xeli_training_preanswer(body: XeliPreanswerIn, request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        return _preanswer(tenant_id, body)

    @router.post("/admin/xeli/operational-table")
    def xeli_operational_table(body: XeliOperationalTableIn, request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        return _operational_error_table(body)

    @router.post("/admin/xeli/training/correct-general")
    def xeli_correct_general(body: XeliCorrectionIn, request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        return _translate_correction(tenant_id, "general", body)

    @router.post("/admin/xeli/training/correct-skill")
    def xeli_correct_skill(body: XeliCorrectionIn, request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        return _translate_correction(tenant_id, "sku", body)

    @router.get("/admin/xeli/knowledge")
    def xeli_knowledge_list(request: Request, sku: str = Query(""), include_inactive: bool = Query(True), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        rows = _knowledge(tenant_id, sku=sku, include_inactive=include_inactive)
        items = [
            {
                "skill_id": row.get("skill_id"),
                "instruction": row.get("instruction"),
                "active": bool(row.get("active", True)),
            }
            for row in rows
        ]
        return {"ok": True, "items": items, "total": len(items), "version": XELI_API_VERSION}

    @router.get("/admin/xeli/knowledge.xlsx")
    def xeli_knowledge_download(request: Request, token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        content = _knowledge_workbook(tenant_id)
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": 'attachment; filename="conocimiento_xeli.xlsx"',
                "X-XelerIA-Filename": "conocimiento_xeli.xlsx",
                "Cache-Control": "no-store",
            },
        )

    @router.post("/admin/xeli/knowledge.xlsx")
    async def xeli_knowledge_upload(request: Request, file: UploadFile = File(...), token: Optional[str] = Query(default=None), x_admin_token: Optional[str] = Header(default=None), x_session_token: Optional[str] = Header(default=None), session_token: Optional[str] = Query(default=None), authorization: Optional[str] = Header(default=None)):
        tenant_id = auth(request, token, x_admin_token, x_session_token, session_token, authorization)
        if not str(file.filename or "").lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")
        try:
            content = await file.read()
        finally:
            await file.close()
        if not content:
            raise HTTPException(status_code=400, detail="El archivo está vacío")
        if len(content) > _KNOWLEDGE_MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=413, detail="El archivo supera el máximo de 15 MB")
        result = _import_knowledge(tenant_id, content)
        if not result.get("ok"):
            raise HTTPException(
                status_code=422,
                detail={
                    "message": "La planilla tiene errores y no se importó.",
                    "errors": result.get("errors") or [],
                },
            )
        return {**result, "version": XELI_API_VERSION}
